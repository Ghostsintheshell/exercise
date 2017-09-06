# -*- coding: utf-8 -*-
"""
Created on Mon Aug 21 22:40:00 2017

@author: Yew
"""

import openpyxl as xl

#创建一个excel对象，打开一个excel文件
wb = xl.load_workbook('ETF_analysis - copy.xlsx')
print('表单名称：',wb.sheetnames)

#选取表单 
ws = wb['Valuation']
"""
#上证50
if ws['E2'].value <= 10.0:
    ws['C2'] = '低估';
elif 10.0 < ws['E2'].value <=40.0:
    ws['C2'] = '合理'
else:
    ws['C2'] = '高估'
"""
#循环错误，需检查
"""
for result in ws['C2':'C5']:
    for ETF in [ws['E2'].value,ws['E3'].value,ws['E4'].value,ws['E5'].value]:
        if ETF <=10.0:
            result = '低估';
        elif 10.0<ETF<=40.0:
            result = '合理';
        else:
            result = '高估'
"""
#Create PE data list
PE_list = []
for n in range(4):
    if ws.cell(row=n+1,column=5).value != None:
        PE_list.append(ws.cell(row=n+1,column=5))

#Valuation Need fix bug

for n in range(len(PE_list)):
    if PE_list[n].value <=10:
        ws.cell(row=n+2,column=3).value = '低估'
    elif 10 < PE_list[n].value <= 40:
        ws.cell(row=n+2,column=3).value = '合理'
    else:
        ws.cell(row=n+2,column=3).value = '高估'
       
"""    
print ws.cell('C2').value,ws.cell('E2').value
print ws.cell('C3').value
print ws.cell('C4').value        
"""
wb.save("ETF_analysis - copy.xlsx")